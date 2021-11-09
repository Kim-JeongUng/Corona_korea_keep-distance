# 소분류 지역별 코로나 확진자
import requests
from bs4 import BeautifulSoup

# 매일 GPS 기록 저장
import pandas as pd
import datetime

import re

import openpyxl

url = requests.get('http://ncov.mohw.go.kr/bdBoardList_Real.do?brdId=1&brdGubun=13&ncvContSeq=&contSeq=&board_id=&gubun=')
CovidCountURL = BeautifulSoup(url.content, "html.parser")

today = datetime.datetime.now()
yesterday = today - datetime.timedelta(1)
today = today.strftime("%Y%m%d")
yesterday = yesterday.strftime("%Y%m%d")

wb = openpyxl.load_workbook('CovidCount.xlsx')



def save():
    # 이부분에 이전 Today자료를 YesterDay시트로 옮겨줘야함
    print("날짜가 바뀌어 데이터를 저장합니다.")
    # today -> yesterday , 오늘 새로운 데이터 추가

    wb.create_sheet(index=0, title=today)
    for i in range(0, 18, 1):
        regionData = CovidCountURL.select("#zone_popup" + str(i) + " > div > table > tbody > tr")

        for rgn in regionData:
            # 콤마 제거
            rgn = rgn.text.replace(',', '')
            data = {
                'date': [today],
                'region': [rgn],
            }
            # 엑셀 저장
            td_sheet = wb[today]
            td_sheet.append([rgn])
            wb.save('CovidCount.xlsx')

def CovidCountSave():
    try:
        # 데이터 파일의 첫번째 인덱스 시트의 값과 오늘 날짜 비교, 다르면 저장
        if int(wb.sheetnames[0]) != int(today):
            save()
            state = 1
        else:
            # 이미 저장된 자료가 있음
            print("이전과 같음")
            state = 0

    except:
        save()
        state = -1

    # state -1 : 오류(파일없음 등) / 0 : 데이터의 변화 없음 저장안함 / 1 : 새로운 데이터를 저장함
    return state


def find(area):
    td_sheet = wb[today]        # 오늘 시트
    ys_sheet = wb[yesterday]    # 전날 시트

    for i in ys_sheet.rows:
        if area in i[0].value:
            ys_cnt = int(re.findall("\d+", i[0].value)[0])
            break

    for i in td_sheet.rows:
        if area in i[0].value:
            td_cnt = int(re.findall("\d+", i[0].value)[0])
            break

    # 오늘 누적 확진자와 어제 누적 확진자의 차이
    cnt_today = td_cnt - ys_cnt
    return cnt_today


def __init__():
    CovidCountSave()


''' 
    서버에선 다음을 실행해야함 
    CovidCountSave()
'''
'''
    사용자는 다음을 실행해야함
    find(area)
    #print(find("원주시"))
    #print(find("춘천시"))
'''
